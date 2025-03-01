from openpyxl import Workbook, load_workbook
import os
import speech_recognition as sr
import pyttsx3
from datetime import datetime
import re

# Initialize the voice engine
listener = sr.Recognizer()
engine = pyttsx3.init()
engine.setProperty('rate', 120)
voices = engine.getProperty('voices')
engine.setProperty('voice', voices[1].id)

def talk(text):
    """Speak the given text."""
    print(text)  # Debugging output
    engine.say(text)
    engine.runAndWait()

def check_exists():
    """Check if the tasks.xlsx file exists."""
    return os.path.exists('tasks.xlsx')

def get_first_nonpopulated(sheet):
    """Find the first non-populated row in the sheet."""
    column = sheet['A']
    for cell in column[::-1]:
        if cell.value is None:
            return cell.row
    return 1

def get_voice_input(prompt_text):
    """Get voice input from the user."""
    while True:
        try:
            talk(prompt_text)
            with sr.Microphone() as source:
                print(prompt_text)
                voice = listener.listen(source)
            input_text = listener.recognize_google(voice, language='en-US').lower()
            print(f"You said: {input_text}")
            return input_text
        except Exception as e:
            print(f"Error: {e}")
            talk("I couldn't understand that. Please try again.")

def get_time_input():
    """Get time input from the user in a valid format."""
    while True:
        try:
            talk("Please tell me the time, like '8 45 PM' or '14 30'.")
            with sr.Microphone() as source:
                print("Listening for time...")
                voice = listener.listen(source)
            
            spoken_time = listener.recognize_google(voice, language='en-US').lower()
            print(f"You said: {spoken_time}")

            # Normalize spoken time format
            spoken_time = spoken_time.replace("a.m.", " am").replace("p.m.", " pm").strip()
            spoken_time = re.sub(r"\s+", " ", spoken_time)  # Remove extra spaces

            # Convert "845 PM" to "8:45 PM" if missing colon
            match = re.match(r"^(\d{1,2})(\d{2})\s?(am|pm)?$", spoken_time)
            if match:
                hour, minute, period = match.groups()
                spoken_time = f"{hour}:{minute} {period}".strip()

            # Try parsing different time formats
            try:
                time_object = datetime.strptime(spoken_time, "%I:%M %p")
            except ValueError:
                try:
                    time_object = datetime.strptime(spoken_time, "%H:%M")
                except ValueError:
                    talk("Invalid time format. Please say it again.")
                    continue

            formatted_time = time_object.strftime("%H:%M")  # Convert to 24-hour format
            print(f"Formatted time: {formatted_time}")
            return formatted_time

        except Exception as e:
            print(f"Error: {e}")
            talk("I couldn't understand the time format. Please try again.")

def main_program():
    """Main program to create tasks and save them in an Excel file."""
    check_exist = check_exists()

    # Get task input
    task = get_voice_input("What is your task?")

    # Get date input (current date)
    date = datetime.now().date().strftime('%Y-%m-%d')

    # Get time input
    time = get_time_input()

    if check_exist:
        workbook = load_workbook('tasks.xlsx')
        sheet = workbook.active
        first_nonpopulated = get_first_nonpopulated(sheet) + 2

        # Write data to the sheet
        sheet.cell(row=first_nonpopulated, column=1, value=task)
        sheet.cell(row=first_nonpopulated, column=2, value=date)
        sheet.cell(row=first_nonpopulated, column=3, value=time)

        workbook.save('tasks.xlsx')
        talk("Task saved successfully!")
    else:
        workbook = Workbook()
        sheet = workbook.active

        # Create headers
        sheet['A1'] = "Task"
        sheet['B1'] = "Date"
        sheet['C1'] = "Time"

        # Write data to the first row
        row = [task, date, time]
        sheet.append(row)
        workbook.save('tasks.xlsx')
        talk("Task saved successfully!")

def extract_time_and_task(command):
    """Extract time and task from the user's voice command."""
    match = re.search(r"(\d{1,2}[:.]?\d{0,2})\s?(am|pm)?\s*(to|for)?\s*(.*)", command)
    if match:
        time_part = match.group(1).replace(".", ":")  # Convert "7.45" to "7:45"
        am_pm = match.group(2) or ""  # Extract AM/PM if spoken
        task = match.group(4).strip()  # Extract the task

        # Convert to 24-hour format
        try:
            if ":" not in time_part:  # Handle "745 PM" case
                hour = int(time_part[:-2]) // 100
                minute = int(time_part[:-2]) % 100
                time_part = f"{hour}:{minute}"

            time_str = f"{time_part} {am_pm}".strip()
            time_object = datetime.strptime(time_str, "%I:%M %p") if am_pm else datetime.strptime(time_part, "%H:%M")
            formatted_time = time_object.strftime("%H:%M")
            talk(task, formatted_time)
            return formatted_time, task
        except ValueError:
            talk("Invalid time format. Please try again.")
            return None, None
    else:
        talk("I couldn't understand the time. Please try again.")
        return None, None
