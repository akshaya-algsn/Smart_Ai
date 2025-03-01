from datetime import datetime
from openpyxl import load_workbook
import pyttsx3
import threading
import time
from plyer import notification

# Initialize TTS engine
engine = pyttsx3.init()
engine.setProperty('rate', 150)
voices = engine.getProperty('voices')
engine.setProperty('voice', voices[1].id)

def talk(text):
    engine.say(text)
    engine.runAndWait()

def read():
    try:
        workbook = load_workbook('tasks.xlsx')
        sheet = workbook.active

        current_date = datetime.now().date().strftime('%Y-%m-%d')
        current_time = datetime.now().time().strftime('%H:%M')

        for row in sheet.iter_rows(min_row=2, values_only=True):
            task, date, time_str = row[:3]

            if date == current_date and time_str == current_time:
                message = f"Reminder! It's time to complete your task: {task}"
                notification.notify(
            title="Notification",
            message="Reminder! It's time to complete your task: {task}",
            timeout=5
    )
                print(message)
                talk(message)
    except FileNotFoundError:
        print("Error: tasks.xlsx file not found.")
    except Exception as e:
        print(f"An error occurred: {e}")

def task_checker():
    """Checks for tasks every 10 seconds without blocking other processes."""
    while True:
        read()
        time.sleep(10)

def start_reading_tasks():
    """Starts the task reminder system in a background thread."""
    task_thread = threading.Thread(target=task_checker, daemon=True)
    task_thread.start()
